"""
Lecture du fichier export TIS quotidien.
Localise les colonnes par NOM DE TAG (ligne 1 de l'export), pas par lettre
de colonne. C'est ce qui rend le script resistant a un export TIS
legerement different (colonnes reordonnees, colonnes en plus...).
"""

import openpyxl


# Mapping nom-metier -> nom exact du tag dans l'export TIS (ligne 1)
# BC1 uniquement pour l'instant : ajouter le mapping BC2 le jour ou BC2
# tourne aussi et qu'il faut le traiter.
TAGS_BC1 = {
    "debit_th20": "Débit total TH20",
    "pw_moteur": "Pw Mot Br",
    "pw_ve62": "Pw VE62",
    "kwh_bc": "Clle_BC1",
    "debit_dos_farine": "Débit Dos Farine",
    "position_volet_alim_four": "Position ON VO18",
    "debit_rejet": "Débit rejet",
    "rejet_vers_sol": "Rejet vers Sol",
    "alim_vers_sol": "Alim vers Sol",
    "niveau_silo": "Niv Silo Farine 1",
    "vitesse_separateur": "Vitesse sép BC1",
}


def trouver_colonnes(feuille, ligne_entete=1):
    """Construit un dict {nom_tag: numero_colonne} en lisant la ligne d'entete."""
    colonnes = {}
    for col in range(1, feuille.max_column + 1):
        valeur = feuille.cell(row=ligne_entete, column=col).value
        if valeur:
            colonnes[str(valeur).strip()] = col
    return colonnes


def lire_export_tis(chemin_fichier, feuille_nom="Donnés TIS BK&BC", ligne_debut=4):
    """
    Lit l'export TIS quotidien et retourne un dict de listes numeriques,
    une entree par variable metier (cf TAGS_BC1), alignees minute par minute.

    ligne_debut=4 reproduit la structure observee : lignes 1-3 = entetes,
    donnee minute par minute a partir de la ligne 4.
    """
    wb = openpyxl.load_workbook(chemin_fichier, data_only=True)
    if feuille_nom not in wb.sheetnames:
        raise ValueError(
            f"Feuille '{feuille_nom}' introuvable. "
            f"Feuilles disponibles : {wb.sheetnames}"
        )
    ws = wb[feuille_nom]
    colonnes_disponibles = trouver_colonnes(ws, ligne_entete=1)

    # Verifie que tous les tags attendus existent bien dans ce fichier
    manquants = [tag for tag in TAGS_BC1.values() if tag not in colonnes_disponibles]
    if manquants:
        raise ValueError(
            f"Tags introuvables dans l'export TIS : {manquants}\n"
            f"L'export a peut-etre change de format. Colonnes trouvees : "
            f"{list(colonnes_disponibles.keys())}"
        )

    donnees = {nom_metier: [] for nom_metier in TAGS_BC1}
    ligne_fin = ws.max_row

    # Date de la journee, lue une seule fois sur la colonne A (colonne 1).
    # Necessaire pour dater automatiquement chaque ligne de l'historique.
    premiere_valeur_date = ws.cell(row=ligne_debut, column=1).value
    if hasattr(premiere_valeur_date, "date"):
        donnees["_date_journee"] = premiere_valeur_date.date()
    else:
        donnees["_date_journee"] = premiere_valeur_date  # valeur brute si pas un datetime

    for row in range(ligne_debut, ligne_fin + 1):
        # On s'arrete si la ligne est vide sur toute la largeur utile
        # (evite de lire les lignes de Sum/Avg en bas de feuille)
        premiere_cellule = ws.cell(row=row, column=colonnes_disponibles[TAGS_BC1["debit_th20"]]).value
        if premiere_cellule is None:
            # une minute d'arret peut avoir une cellule vide -> continuer,
            # mais s'arreter si TOUTE la ligne est vide (fin de journee)
            toute_vide = all(
                ws.cell(row=row, column=col).value is None
                for col in colonnes_disponibles.values()
            )
            if toute_vide:
                break

        for nom_metier, tag in TAGS_BC1.items():
            col = colonnes_disponibles[tag]
            valeur = ws.cell(row=row, column=col).value
            donnees[nom_metier].append(valeur)

    return donnees